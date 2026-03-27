# ============================================================
# WEB SCRAPER PRO v1.0
# Scrape any website and export data to Excel
# Clean dark PyQt5 GUI
# by Areeb
# ============================================================

import sys
import threading
import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog, QFrame,
    QProgressBar, QTextEdit, QSplitter, QStatusBar
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QColor, QPalette

# ============================================================
# THEME
# ============================================================
STYLE = """
QMainWindow, QWidget {
    background-color: #0f1923;
    color: #e8f0fe;
    font-family: 'Segoe UI';
    font-size: 10pt;
}
QLineEdit, QComboBox, QTextEdit {
    background-color: #1a2634;
    color: #e8f0fe;
    border: 1px solid #243447;
    border-radius: 4px;
    padding: 6px 10px;
    font-size: 10pt;
}
QLineEdit:focus, QComboBox:focus {
    border: 1px solid #2196f3;
}
QComboBox::drop-down {
    border: none;
    padding-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #1a2634;
    color: #e8f0fe;
    selection-background-color: #2196f3;
    border: 1px solid #243447;
}
QPushButton {
    background-color: #2196f3;
    color: #ffffff;
    border: none;
    border-radius: 4px;
    padding: 8px 18px;
    font-weight: bold;
    font-size: 10pt;
}
QPushButton:hover { background-color: #1e88e5; }
QPushButton:pressed { background-color: #1565c0; }
QPushButton:disabled { background-color: #243447; color: #90a4ae; }
QPushButton#clearBtn {
    background-color: #243447;
    color: #90a4ae;
}
QPushButton#clearBtn:hover { background-color: #2d4057; color: #e8f0fe; }
QPushButton#exportBtn {
    background-color: #00897b;
}
QPushButton#exportBtn:hover { background-color: #00796b; }
QPushButton#exportBtn:disabled { background-color: #243447; color: #90a4ae; }
QTableWidget {
    background-color: #1a2634;
    color: #e8f0fe;
    border: 1px solid #243447;
    border-radius: 4px;
    gridline-color: #243447;
    font-size: 9pt;
}
QTableWidget::item { padding: 6px; }
QTableWidget::item:selected {
    background-color: #1565c0;
    color: #ffffff;
}
QHeaderView::section {
    background-color: #243447;
    color: #90a4ae;
    padding: 6px;
    border: none;
    border-right: 1px solid #0f1923;
    font-weight: bold;
    font-size: 9pt;
}
QScrollBar:vertical {
    background: #1a2634;
    width: 8px;
    border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #243447;
    border-radius: 4px;
}
QScrollBar::handle:vertical:hover { background: #2196f3; }
QScrollBar:horizontal {
    background: #1a2634;
    height: 8px;
    border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #243447;
    border-radius: 4px;
}
QProgressBar {
    background-color: #1a2634;
    border: 1px solid #243447;
    border-radius: 4px;
    height: 6px;
    text-align: center;
}
QProgressBar::chunk {
    background-color: #2196f3;
    border-radius: 4px;
}
QStatusBar {
    background-color: #0f1923;
    color: #90a4ae;
    font-size: 9pt;
    border-top: 1px solid #243447;
}
QSplitter::handle { background-color: #243447; }
QFrame#card {
    background-color: #1a2634;
    border: 1px solid #243447;
    border-radius: 6px;
}
QLabel#heading {
    font-size: 13pt;
    font-weight: bold;
    color: #e8f0fe;
}
QLabel#sub {
    font-size: 9pt;
    color: #90a4ae;
}
QLabel#sectionTitle {
    font-size: 9pt;
    font-weight: bold;
    color: #90a4ae;
    text-transform: uppercase;
    letter-spacing: 1px;
}
QLabel#countBadge {
    background-color: #2196f3;
    color: #ffffff;
    border-radius: 10px;
    padding: 2px 10px;
    font-size: 9pt;
    font-weight: bold;
}
QTextEdit {
    font-family: 'Consolas';
    font-size: 9pt;
}
"""

# ============================================================
# SCRAPER WORKER THREAD
# ============================================================
class ScraperWorker(QThread):
    result  = pyqtSignal(list, list)   # headers, rows
    error   = pyqtSignal(str)
    log     = pyqtSignal(str, str)     # message, color
    done    = pyqtSignal()

    def __init__(self, url, mode, custom_selector=""):
        super().__init__()
        self.url             = url
        self.mode            = mode
        self.custom_selector = custom_selector

    def run(self):
        try:
            self.log.emit(f"Connecting to {self.url} ...", "#90a4ae")
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                )
            }
            response = requests.get(self.url, headers=headers, timeout=15)
            response.raise_for_status()
            self.log.emit(f"Got response: {response.status_code} OK", "#00e676")

            soup = BeautifulSoup(response.text, "html.parser")
            self.log.emit(f"Parsing page with mode: {self.mode} ...", "#90a4ae")

            if self.mode == "Links":
                cols, rows = self._scrape_links(soup)
            elif self.mode == "Headings":
                cols, rows = self._scrape_headings(soup)
            elif self.mode == "Images":
                cols, rows = self._scrape_images(soup)
            elif self.mode == "Tables":
                cols, rows = self._scrape_tables(soup)
            elif self.mode == "Custom Selector":
                cols, rows = self._scrape_custom(soup)
            else:
                cols, rows = [], []

            self.log.emit(f"Done! Found {len(rows)} result(s).", "#2196f3")
            self.result.emit(cols, rows)

        except requests.exceptions.ConnectionError:
            self.error.emit("Connection failed. Check the URL and your internet.")
        except requests.exceptions.Timeout:
            self.error.emit("Request timed out. The site took too long to respond.")
        except requests.exceptions.HTTPError as e:
            self.error.emit(f"HTTP error: {e}")
        except Exception as e:
            self.error.emit(f"Unexpected error: {e}")
        finally:
            self.done.emit()

    def _scrape_links(self, soup):
        cols = ["#", "Text", "URL", "Type"]
        rows = []
        for i, tag in enumerate(soup.find_all("a", href=True), 1):
            href = tag["href"].strip()
            text = tag.get_text(strip=True) or "(no text)"
            link_type = "External" if href.startswith("http") else "Internal"
            rows.append([str(i), text[:120], href[:200], link_type])
        return cols, rows

    def _scrape_headings(self, soup):
        cols = ["#", "Tag", "Text"]
        rows = []
        for i, tag in enumerate(soup.find_all(["h1","h2","h3","h4","h5","h6"]), 1):
            rows.append([str(i), tag.name.upper(), tag.get_text(strip=True)[:200]])
        return cols, rows

    def _scrape_images(self, soup):
        cols = ["#", "Alt Text", "Source URL", "Width", "Height"]
        rows = []
        for i, tag in enumerate(soup.find_all("img"), 1):
            src    = tag.get("src", "(none)")
            alt    = tag.get("alt", "(no alt)")
            width  = tag.get("width", "—")
            height = tag.get("height", "—")
            rows.append([str(i), alt[:100], src[:200], width, height])
        return cols, rows

    def _scrape_tables(self, soup):
        tables = soup.find_all("table")
        if not tables:
            return ["Info"], [["No tables found on this page."]]
        table   = tables[0]
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        if not headers:
            headers = [f"Col {i+1}" for i in range(
                len(table.find("tr").find_all("td")))]
        cols = ["#"] + headers
        rows = []
        for i, tr in enumerate(table.find_all("tr")[1:], 1):
            cells = [td.get_text(strip=True)[:100] for td in tr.find_all("td")]
            if cells:
                rows.append([str(i)] + cells)
        return cols, rows

    def _scrape_custom(self, soup):
        if not self.custom_selector.strip():
            return ["Info"], [["Enter a CSS selector above."]]
        elements = soup.select(self.custom_selector.strip())
        cols = ["#", "Tag", "Text", "HTML"]
        rows = []
        for i, el in enumerate(elements, 1):
            rows.append([
                str(i),
                el.name,
                el.get_text(strip=True)[:150],
                str(el)[:200]
            ])
        return cols, rows


# ============================================================
# MAIN WINDOW
# ============================================================
class WebScraperApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Web Scraper Pro")
        self.setMinimumSize(900, 680)
        self.resize(1000, 720)
        self.setStyleSheet(STYLE)

        self._cols = []
        self._rows = []
        self._worker = None

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(20, 16, 20, 8)
        layout.setSpacing(14)

        # HEADER
        header = QWidget()
        hl = QHBoxLayout(header)
        hl.setContentsMargins(0, 0, 0, 0)

        title_block = QVBoxLayout()
        title_lbl = QLabel("Web Scraper Pro")
        title_lbl.setObjectName("heading")
        sub_lbl = QLabel("Extract data from any website and export to Excel")
        sub_lbl.setObjectName("sub")
        title_block.addWidget(title_lbl)
        title_block.addWidget(sub_lbl)
        title_block.setSpacing(2)

        hl.addLayout(title_block)
        hl.addStretch()

        self.count_badge = QLabel("0 results")
        self.count_badge.setObjectName("countBadge")
        self.count_badge.setVisible(False)
        hl.addWidget(self.count_badge)

        layout.addWidget(header)

        # URL BAR CARD
        url_card = QFrame()
        url_card.setObjectName("card")
        url_layout = QVBoxLayout(url_card)
        url_layout.setContentsMargins(16, 14, 16, 14)
        url_layout.setSpacing(10)

        url_row = QHBoxLayout()
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("https://example.com")
        self.url_input.returnPressed.connect(self._start_scrape)
        url_row.addWidget(self.url_input)

        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["Links", "Headings", "Images", "Tables", "Custom Selector"])
        self.mode_combo.setFixedWidth(160)
        self.mode_combo.currentTextChanged.connect(self._on_mode_change)
        url_row.addWidget(self.mode_combo)

        self.scrape_btn = QPushButton("⚡  Scrape")
        self.scrape_btn.setFixedWidth(110)
        self.scrape_btn.clicked.connect(self._start_scrape)
        url_row.addWidget(self.scrape_btn)

        url_layout.addLayout(url_row)

        # Custom selector row (hidden by default)
        self.custom_row = QHBoxLayout()
        custom_lbl = QLabel("CSS Selector:")
        custom_lbl.setObjectName("sectionTitle")
        custom_lbl.setFixedWidth(100)
        self.custom_input = QLineEdit()
        self.custom_input.setPlaceholderText("e.g.  div.product-title  or  table.data  or  span.price")
        self.custom_row.addWidget(custom_lbl)
        self.custom_row.addWidget(self.custom_input)
        self.custom_widget = QWidget()
        self.custom_widget.setLayout(self.custom_row)
        self.custom_widget.setVisible(False)
        url_layout.addWidget(self.custom_widget)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setVisible(False)
        self.progress.setFixedHeight(5)
        url_layout.addWidget(self.progress)

        layout.addWidget(url_card)

        # SPLITTER — table + log
        splitter = QSplitter(Qt.Vertical)
        splitter.setHandleWidth(6)

        # RESULTS TABLE
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(6)

        tbl_header = QHBoxLayout()
        tbl_lbl = QLabel("RESULTS")
        tbl_lbl.setObjectName("sectionTitle")
        tbl_header.addWidget(tbl_lbl)
        tbl_header.addStretch()

        self.export_btn = QPushButton("💾  Export to Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export_excel)
        tbl_header.addWidget(self.export_btn)

        self.clear_btn = QPushButton("✕  Clear")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.clicked.connect(self._clear)
        tbl_header.addWidget(self.clear_btn)

        table_layout.addLayout(tbl_header)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            "QTableWidget { alternate-background-color: #162230; }")
        table_layout.addWidget(self.table)

        splitter.addWidget(table_widget)

        # LOG
        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.setSpacing(4)

        log_lbl = QLabel("LOG")
        log_lbl.setObjectName("sectionTitle")
        log_layout.addWidget(log_lbl)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFixedHeight(90)
        log_layout.addWidget(self.log_box)

        splitter.addWidget(log_widget)
        splitter.setSizes([480, 110])

        layout.addWidget(splitter)

        # STATUS BAR
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("Ready — enter a URL and click Scrape")

    # ── SLOTS ────────────────────────────────────────────────
    def _on_mode_change(self, mode):
        self.custom_widget.setVisible(mode == "Custom Selector")

    def _start_scrape(self):
        url = self.url_input.text().strip()
        if not url:
            self.status.showMessage("⚠  Please enter a URL first.")
            return
        if not url.startswith("http"):
            url = "https://" + url
            self.url_input.setText(url)

        self.scrape_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.count_badge.setVisible(False)
        self._clear_table()
        self.log_box.clear()
        self.status.showMessage("Scraping...")

        mode   = self.mode_combo.currentText()
        custom = self.custom_input.text() if mode == "Custom Selector" else ""

        self._worker = ScraperWorker(url, mode, custom)
        self._worker.result.connect(self._on_result)
        self._worker.error.connect(self._on_error)
        self._worker.log.connect(self._on_log)
        self._worker.done.connect(self._on_done)
        self._worker.start()

    def _on_result(self, cols, rows):
        self._cols = cols
        self._rows = rows
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                item = QTableWidgetItem(val)
                self.table.setItem(r, c, item)

        self.table.resizeColumnsToContents()
        self.table.horizontalHeader().setStretchLastSection(True)

        self.count_badge.setText(f"{len(rows)} results")
        self.count_badge.setVisible(True)
        self.export_btn.setEnabled(len(rows) > 0)
        self.status.showMessage(f"Done — {len(rows)} result(s) found.")

    def _on_error(self, msg):
        self._log(f"ERROR: {msg}", "#ff5252")
        self.status.showMessage(f"Error: {msg}")

    def _on_log(self, msg, color):
        self._log(msg, color)

    def _on_done(self):
        self.scrape_btn.setEnabled(True)
        self.progress.setVisible(False)

    def _log(self, msg, color="#90a4ae"):
        self.log_box.append(f'<span style="color:{color};">{msg}</span>')

    def _clear_table(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._cols = []
        self._rows = []

    def _clear(self):
        self._clear_table()
        self.log_box.clear()
        self.count_badge.setVisible(False)
        self.export_btn.setEnabled(False)
        self.status.showMessage("Cleared.")

    def _export_excel(self):
        if not self._rows:
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default   = f"scraped_data_{timestamp}.xlsx"
        path, _   = QFileDialog.getSaveFileName(
            self, "Save Excel File", default,
            "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"

        # Header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="2196F3")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for c, col in enumerate(self._cols, 1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r, row in enumerate(self._rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(path)
        self.status.showMessage(f"Exported to {path}")
        self._log(f"Saved: {path}", "#00e676")


# ── RUN ─────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = WebScraperApp()
    window.show()
    sys.exit(app.exec_())
